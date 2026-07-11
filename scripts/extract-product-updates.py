#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract-product-updates.py — 一次性脚本：把 Excel 里的产品更新清单解析成
结构化 JSON，供 `scripts/import-product-updates.mjs`（Prisma Client）入库。

用法：
    python scripts/extract-product-updates.py

输出：
    scripts/product-updates.json  — 解析后的更新清单（含 sku/title/slug/keywords/urls）
    scripts/redirects.json        — 仅 URL 发生变化的产品，旧->新 301 重定向规则

列说明（sheet "Product Title URL"）：
    SKU           产品 SKU（唯一标识，用于匹配 DB）
    修改后的标题    更新 product.name.en（非空才更新）
    修改后的URL     完整 URL，提取 slug 后更新 product.slug（非空才更新）
    修改关键词      更新 product.seoKeywords.en（逗号分隔字符串，非空才更新）
    原网站URL / 修改后的URL  用于生成 301 重定向（二者都非空且不同才生成）

slug 提取规则（与 src/lib/slug.ts 的 normalizeSlug 保持一致）：
    - 去掉 `https://www.wstoolcabinet.com/en/` 域名前缀；**保留 `.html` 后缀**（URL 形态必须与表格一致）
    - 例如 `.../en/products/automated-tool-storage-system.html`
      -> `products/automated-tool-storage-system.html`
    - 例如 `.../en/products/automated-tool-storage-system.html` -> `products/automated-tool-storage-system.html`
    - 叶子段小写、空格转连字符、去除非法字符、折叠多余连字符
"""

import json
import os
import re

import openpyxl

# ----------------------------------------------------------------------------
# 路径
# ----------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
EXCEL_PATH = r"C:/Users/Administrator/Desktop/产品标题+URL.xlsx"
SHEET_NAME = "Product Title URL"
OUT_UPDATES = os.path.join(HERE, "product-updates.json")
OUT_REDIRECTS = os.path.join(HERE, "redirects.json")

# ----------------------------------------------------------------------------
# slug 规范化（移植自 src/lib/slug.ts 的 normalizeSlug / normalizeLeaf）
# ----------------------------------------------------------------------------

def normalize_leaf(leaf: str) -> str:
    leaf = (leaf or "").lower()
    leaf = re.sub(r"\s+", "-", leaf)
    leaf = re.sub(r"[^a-z0-9.\-]", "", leaf)
    leaf = re.sub(r"\.{2,}", ".", leaf)
    leaf = re.sub(r"-{2,}", "-", leaf)
    leaf = re.sub(r"^-+", "", leaf)
    leaf = re.sub(r"-+$", "", leaf)
    return leaf


def normalize_slug(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return ""
    last_slash = value.rfind("/")
    if last_slash >= 0:
        directory = value[: last_slash + 1]
        leaf = value[last_slash + 1 :]
        return f"{directory}{normalize_leaf(leaf)}"
    return normalize_leaf(value)


def slug_from_url(url: str):
    """从完整 URL 提取规范化后的 slug（去掉域名前缀，保留 .html 后缀）。"""
    if not url:
        return None
    url = str(url).strip()
    # 去掉域名前缀：取 /en/ 之后的部分
    parts = url.split("/en/", 1)
    if len(parts) < 2:
        # 兜底：如果没有 /en/，尝试取最后一个路径段
        path = url.split("?", 1)[0]
        seg = path.rstrip("/").split("/")[-1]
        if not seg:
            return None
        return normalize_slug(seg)
    s = parts[1].split("?", 1)[0]  # 去掉 query
    s = s.strip("/")
    if not s:
        return None
    return normalize_slug(s)


def clean(v):
    """Excel 单元格 -> 字符串；None / 空 -> ''。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' 不存在。可用 sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def col(name):
        if name not in idx:
            raise SystemExit(f"Excel 缺少列: {name}（实际列: {list(idx.keys())}）")
        return idx[name]

    c_sku = col("SKU")
    c_new_title = col("修改后的标题")
    c_new_url = col("修改后的URL")
    c_new_kw = col("修改关键词")
    c_old_url = col("原网站URL")

    updates = []
    redirects = []
    skipped = 0

    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        sku = clean(r[c_sku])
        if not sku:
            skipped += 1
            continue

        new_title = clean(r[c_new_title])
        new_url = clean(r[c_new_url])
        new_kw = clean(r[c_new_kw])
        old_url = clean(r[c_old_url])

        entry = {
            "sku": sku,
            "title": new_title,        # 非空才更新 name.en
            "slug": slug_from_url(new_url) if new_url else "",  # 非空才更新 slug
            "keywords": new_kw,        # 非空才更新 seoKeywords.en
            "oldUrl": old_url,
            "newUrl": new_url,
        }
        updates.append(entry)

        # 仅当 URL 确实变化时才生成 301 重定向
        if old_url and new_url and old_url.strip() != new_url.strip():
            old_slug = slug_from_url(old_url)
            new_slug = slug_from_url(new_url)
            if old_slug and new_slug and old_slug != new_slug:
                redirects.append(
                    {
                        "source": f"/en/{old_slug}",
                        "destination": f"/en/{new_slug}",
                        "permanent": True,
                    }
                )

    with open(OUT_UPDATES, "w", encoding="utf-8") as f:
        json.dump(updates, f, ensure_ascii=False, indent=2)
    with open(OUT_REDIRECTS, "w", encoding="utf-8") as f:
        json.dump(redirects, f, ensure_ascii=False, indent=2)

    print(f"[OK] 解析 {len(updates)} 条更新（跳过空 SKU 行 {skipped}）")
    print(f"[OK] 生成 {len(redirects)} 条 301 重定向 -> {os.path.relpath(OUT_REDIRECTS, ROOT)}")
    print(f"[OK] 更新清单 -> {os.path.relpath(OUT_UPDATES, ROOT)}")
    print("\n--- 重定向规则预览 ---")
    for rd in redirects:
        print(f"  {rd['source']}  ->  {rd['destination']}  (permanent={rd['permanent']})")


if __name__ == "__main__":
    main()
